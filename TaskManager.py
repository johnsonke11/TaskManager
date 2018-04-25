import os
from PyQt5 import QtCore, QtGui, QtWidgets
import taskdesign
import openpyxl
import datetime
import sys
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QMessageBox, QTableWidgetItem,QVBoxLayout, QTableWidget

class tasks(object):
    def __init__(self,task=None,person=None,time=None):
        self.person = str(person)
        self.task = str(task)
        self.time = str(time)
    def get_person(self):
        return self.person
    def set_person(self, x):
        self.person = x
    def get_task(self):
        return self.task
    def get_time(self):
        return self.time


WORKBOOK_NAME = "tasks.xlsx"
if (os.path.exists(WORKBOOK_NAME) == False):
  wb = openpyxl.Workbook()
  sheet = wb.active
  sheet.title= 'Sheet1'
  wb.save(WORKBOOK_NAME)
wb = openpyxl.load_workbook(WORKBOOK_NAME)
tasklists = []
def ifexists(name, tasklists):
  i = 0
  for tasks in tasklists:
    if name in tasks.task:
      return i;
    i+=1
  return -1;
def findstring(x):
    for i in range(0,len(x)):
        if (x[i] == "*"):
            return x[0:i+1]
        i+=1
    return -1;
class TaskManager(QtWidgets.QMainWindow, taskdesign.Ui_MainWindow):
    def __init__(self, parent=None):
        super(TaskManager, self).__init__(parent)
        self.setupUi(self)
        self.setupapp()
    def setupapp(self):
        self.Tasklist.clear()
        self.SelectPerson.addItem("Scott")
        self.SelectPerson.addItem("Sonya")
        self.SelectPerson.addItem("Jimmy")
        self.SelectPerson.addItem("Jarvis")
        sheet = wb['Sheet1']
        if (sheet.max_row > 1):
            for i in range(2,sheet.max_row+1):
              newtask= tasks(sheet.cell(row=i, column=1).value,sheet.cell(row=i, column=2).value,sheet.cell(row=i, column=3).value)
              tasklists.append(newtask)
              x = newtask.task + " " + newtask.person + " " + newtask.time
              self.Tasklist.addItem(x)
        self.AddTask.clicked.connect(self.addTask)
        self.Removetask.clicked.connect(self.removeTask)
        self.ChangePerson.clicked.connect(self.changePerson)
    def addTask(self):
        dt = datetime.datetime.now()
        time = dt.strftime('%Y/%m/%d %H:%M:%S')
        task = self.TaskName.text() + "*"
        person = str(self.SelectPerson.currentText())
        newtask = tasks(task,person,time)
        tasklists.append(newtask)
        sheet = wb['Sheet1']
        i = sheet.max_row + 1
        sheet['A' + str(i)]  = task 
        sheet['B'+ str(i)] = person
        sheet['C' + str(i)] = time
        wb.save(WORKBOOK_NAME)
        x = newtask.task + " " + newtask.person + " " + newtask.time
        self.Tasklist.addItem(x)
    def removeTask(self):
        item = self.Tasklist.currentItem()
        text= str(item.text())
        x = findstring(text)
        index = ifexists(x,tasklists)
        items = tasklists[index]
        y = self.Tasklist.takeItem(index)
        tasklists.remove(items)
        i = 2
        sheet = wb['Sheet1']
        wb.remove(sheet)
        wb.create_sheet('Sheet1')
        sheet = wb['Sheet1']
        sheet['A1'] = 'Task'
        sheet['B1'] = 'Person'
        sheet['C1'] = 'Time'
        for taskitem in tasklists:
            print(taskitem.task)
            sheet['A' + str(i)] = taskitem.task
            sheet['B' + str(i)] = taskitem.person
            sheet['C' + str(i)] = taskitem.time
            i+=1
        wb.save(WORKBOOK_NAME)
    def changePerson(self):
        item = self.Tasklist.currentItem()
        text= str(item.text())
        x = findstring(text)
        index = ifexists(x,tasklists)
        items = tasklists[index]
        items.person = str(self.SelectPerson.currentText())
        sheet = wb['Sheet1']
        sheet['B' + str(index+2)] = items.person
        x = items.task + " " + items.person + " " + items.time
        item.setText(x)
        wb.save(WORKBOOK_NAME)
def main():
    app = QtWidgets.QApplication(sys.argv)
    form = TaskManager()
    form.show()
    app.exec_()

if __name__ == '__main__':
    main()
